"""
Purchase Transaction Endpoints
API endpoints for purchase transaction management (SuperAdmin module)
"""

from fastapi import APIRouter, Depends, HTTPException, status, Query, Response
from sqlalchemy.orm import Session
from typing import List, Optional, Dict, Any
import logging
from datetime import datetime
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

from app.core.database import get_db
from app.api.v1.deps import get_current_superadmin, get_current_user
from app.models.user import User
from app.schemas.purchase_transaction import (
    PurchaseTransactionCreate,
    PurchaseTransactionUpdate,
    PurchaseTransactionResponse,
    PaymentStatusUpdate,
    TransactionListResponse,
    TransactionSummary,
    TransactionAnalytics,
    TransactionFilters
)
from app.services.purchase_transaction_service import PurchaseTransactionService, TransactionAnalyticsService

logger = logging.getLogger(__name__)

router = APIRouter()


# ============================================================================
# SUPERADMIN PURCHASE TRANSACTION MANAGEMENT
# ============================================================================

@router.get("/", response_model=List[TransactionListResponse])
def get_all_transactions(
    skip: int = Query(0, ge=0, description="Pagination offset"),
    limit: int = Query(100, ge=1, le=500, description="Maximum records"),
    status: Optional[str] = Query(None, description="Filter by payment status"),
    payment_method: Optional[str] = Query(None, description="Filter by payment method"),
    date_from: Optional[datetime] = Query(None, description="Filter from date"),
    date_to: Optional[datetime] = Query(None, description="Filter to date"),
    search: Optional[str] = Query(None, description="Search term"),
    sort_by: Optional[str] = Query("transaction_date", description="Sort field"),
    sort_order: Optional[str] = Query("desc", description="Sort order (asc/desc)"),
    db: Session = Depends(get_db),
    superadmin: User = Depends(get_current_superadmin)
):
    """
    Get all purchase transactions for superadmin dashboard.
    
    Returns transactions in the format expected by frontend:
    - id, company, email, date, amount, method, status, logo
    """
    try:
        service = PurchaseTransactionService(db)
        
        # Create filters object
        filters = TransactionFilters(
            status=status,
            payment_method=payment_method,
            date_from=date_from,
            date_to=date_to,
            search_term=search,
            sort_by=sort_by,
            sort_order=sort_order
        )
        
        transactions = service.get_all_transactions_for_frontend(
            skip=skip,
            limit=limit,
            filters=filters
        )
        
        logger.info(f"Superadmin {superadmin.email} retrieved {len(transactions)} transactions")
        return transactions
        
    except Exception as e:
        logger.error(f"Error fetching transactions: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error fetching transactions: {str(e)}"
        )


@router.get("/summary", response_model=TransactionSummary)
def get_transaction_summary(
    db: Session = Depends(get_db),
    superadmin: User = Depends(get_current_superadmin)
):
    """
    Get transaction statistics summary for dashboard.
    """
    try:
        service = PurchaseTransactionService(db)
        stats = service.get_transaction_stats()
        
        summary = TransactionSummary(
            total_transactions=stats["total_transactions"],
            total_revenue=stats["total_revenue"],
            paid_transactions=stats["paid_transactions"],
            unpaid_transactions=stats["unpaid_transactions"],
            pending_transactions=stats["pending_transactions"],
            failed_transactions=stats["failed_transactions"],
            average_transaction_amount=stats["average_transaction_amount"]
        )
        
        logger.info(f"Transaction summary retrieved by {superadmin.email}")
        return summary
        
    except Exception as e:
        logger.error(f"Error fetching transaction summary: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error fetching transaction summary: {str(e)}"
        )


@router.get("/{transaction_id}", response_model=PurchaseTransactionResponse)
def get_transaction(
    transaction_id: int,
    db: Session = Depends(get_db),
    superadmin: User = Depends(get_current_superadmin)
):
    """
    Get transaction details by ID.
    """
    try:
        service = PurchaseTransactionService(db)
        transaction = service.get_transaction_by_id(transaction_id)
        
        if not transaction:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Transaction not found"
            )
        
        # Format response for frontend
        response = PurchaseTransactionResponse.model_validate(transaction)
        if transaction.business:
            response.company = transaction.business.business_name
            response.email = transaction.business.owner.email if transaction.business.owner else "N/A"
        
        response.date = transaction.transaction_date.strftime("%d %b %Y")
        response.amount = float(transaction.total_amount)
        response.method = transaction.payment_method
        response.status = transaction.payment_status
        response.logo = f"/assets/img/icons/{transaction.plan_name.lower().replace(' ', '-')}-icon.svg"
        
        # Invoice details
        response.from_info = {
            "name": transaction.invoice_from_name,
            "address": transaction.invoice_from_address or "456 Green St, Hyderabad",
            "email": transaction.invoice_from_email or "info@dcm.com"
        }
        response.to_info = {
            "name": transaction.invoice_to_name,
            "address": transaction.invoice_to_address or "Business Address",
            "email": transaction.invoice_to_email
        }
        
        logger.info(f"Transaction {transaction_id} retrieved by {superadmin.email}")
        return response
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching transaction {transaction_id}: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error fetching transaction: {str(e)}"
        )


@router.get("/invoice/{invoice_id}", response_model=PurchaseTransactionResponse)
def get_transaction_by_invoice_id(
    invoice_id: str,
    db: Session = Depends(get_db),
    superadmin: User = Depends(get_current_superadmin)
):
    """
    Get transaction details by invoice ID.
    """
    try:
        service = PurchaseTransactionService(db)
        transaction = service.get_transaction_by_invoice_id(invoice_id)
        
        if not transaction:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Transaction not found"
            )
        
        logger.info(f"Transaction {invoice_id} retrieved by {superadmin.email}")
        return PurchaseTransactionResponse.model_validate(transaction)
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching transaction {invoice_id}: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error fetching transaction: {str(e)}"
        )


@router.post("/", response_model=PurchaseTransactionResponse, status_code=status.HTTP_201_CREATED)
def create_transaction(
    transaction_data: PurchaseTransactionCreate,
    db: Session = Depends(get_db),
    superadmin: User = Depends(get_current_superadmin)
):
    """
    Create a new purchase transaction.
    """
    try:
        service = PurchaseTransactionService(db)
        
        # Generate invoice ID if not provided
        if not transaction_data.invoice_id:
            transaction_data.invoice_id = service.generate_invoice_id()
        
        transaction = service.create_transaction(transaction_data)
        
        logger.info(f"Transaction created by {superadmin.email}: {transaction.invoice_id}")
        return PurchaseTransactionResponse.model_validate(transaction)
        
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    except Exception as e:
        logger.error(f"Error creating transaction: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error creating transaction: {str(e)}"
        )


@router.put("/{transaction_id}", response_model=PurchaseTransactionResponse)
def update_transaction(
    transaction_id: int,
    update_data: PurchaseTransactionUpdate,
    db: Session = Depends(get_db),
    superadmin: User = Depends(get_current_superadmin)
):
    """
    Update a purchase transaction.
    """
    try:
        service = PurchaseTransactionService(db)
        transaction = service.update_transaction(transaction_id, update_data)
        
        if not transaction:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Transaction not found"
            )
        
        logger.info(f"Transaction {transaction_id} updated by {superadmin.email}")
        return PurchaseTransactionResponse.model_validate(transaction)
        
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating transaction {transaction_id}: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error updating transaction: {str(e)}"
        )


@router.patch("/{transaction_id}/payment-status", response_model=PurchaseTransactionResponse)
def update_payment_status(
    transaction_id: str,  # Changed to str to accept invoice_id
    payment_data: PaymentStatusUpdate,
    db: Session = Depends(get_db),
    superadmin: User = Depends(get_current_superadmin)
):
    """
    Update payment status for a transaction.
    Accepts either database ID (integer) or invoice_id (string).
    """
    try:
        service = PurchaseTransactionService(db)
        
        # Try to convert to int first (database ID)
        try:
            db_transaction_id = int(transaction_id)
            transaction = service.get_transaction_by_id(db_transaction_id)
        except ValueError:
            # If conversion fails, treat as invoice_id
            transaction = service.get_transaction_by_invoice_id(transaction_id)
            if transaction:
                db_transaction_id = transaction.id
            else:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail="Transaction not found"
                )
        
        if not transaction:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Transaction not found"
            )
        
        # Update payment status
        updated_transaction = service.update_payment_status(
            db_transaction_id,
            payment_data.payment_status,
            payment_data.payment_reference,
            payment_data.gateway_response
        )
        
        if not updated_transaction:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Failed to update transaction"
            )
        
        logger.info(f"Payment status updated for transaction {transaction_id} by {superadmin.email}")
        return PurchaseTransactionResponse.model_validate(updated_transaction)
        
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating payment status for transaction {transaction_id}: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error updating payment status: {str(e)}"
        )


@router.delete("/{transaction_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_transaction(
    transaction_id: int,
    db: Session = Depends(get_db),
    superadmin: User = Depends(get_current_superadmin)
):
    """
    Delete a purchase transaction.
    """
    try:
        service = PurchaseTransactionService(db)
        success = service.delete_transaction(transaction_id)
        
        if not success:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Transaction not found"
            )
        
        logger.info(f"Transaction {transaction_id} deleted by {superadmin.email}")
        
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting transaction {transaction_id}: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error deleting transaction: {str(e)}"
        )


# ============================================================================
# TRANSACTION ANALYTICS
# ============================================================================

@router.get("/analytics/dashboard", response_model=Dict[str, Any])
def get_transaction_analytics(
    db: Session = Depends(get_db),
    superadmin: User = Depends(get_current_superadmin)
):
    """
    Get comprehensive transaction analytics for dashboard.
    """
    try:
        analytics_service = TransactionAnalyticsService(db)
        analytics = analytics_service.get_dashboard_analytics()
        
        logger.info(f"Transaction analytics retrieved by {superadmin.email}")
        return analytics
        
    except Exception as e:
        logger.error(f"Error fetching transaction analytics: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error fetching transaction analytics: {str(e)}"
        )


@router.get("/analytics/revenue", response_model=List[Dict[str, Any]])
def get_revenue_analytics(
    months: int = Query(12, ge=1, le=24, description="Number of months"),
    db: Session = Depends(get_db),
    superadmin: User = Depends(get_current_superadmin)
):
    """
    Get revenue analytics by month.
    """
    try:
        service = PurchaseTransactionService(db)
        revenue_data = service.get_monthly_revenue(months)
        
        logger.info(f"Revenue analytics retrieved by {superadmin.email}")
        return revenue_data
        
    except Exception as e:
        logger.error(f"Error fetching revenue analytics: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error fetching revenue analytics: {str(e)}"
        )


@router.get("/analytics/payment-methods", response_model=List[Dict[str, Any]])
def get_payment_method_analytics(
    db: Session = Depends(get_db),
    superadmin: User = Depends(get_current_superadmin)
):
    """
    Get payment method analytics.
    """
    try:
        service = PurchaseTransactionService(db)
        payment_stats = service.get_payment_method_stats()
        
        logger.info(f"Payment method analytics retrieved by {superadmin.email}")
        return payment_stats
        
    except Exception as e:
        logger.error(f"Error fetching payment method analytics: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error fetching payment method analytics: {str(e)}"
        )


@router.get("/analytics/top-customers", response_model=List[Dict[str, Any]])
def get_top_customers(
    limit: int = Query(10, ge=1, le=50, description="Number of top customers"),
    db: Session = Depends(get_db),
    superadmin: User = Depends(get_current_superadmin)
):
    """
    Get top customers by transaction volume.
    """
    try:
        service = PurchaseTransactionService(db)
        top_customers = service.get_top_customers(limit)
        
        logger.info(f"Top customers analytics retrieved by {superadmin.email}")
        return top_customers
        
    except Exception as e:
        logger.error(f"Error fetching top customers: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error fetching top customers: {str(e)}"
        )


# ============================================================================
# PAYMENT MANAGEMENT
# ============================================================================

@router.get("/pending", response_model=List[PurchaseTransactionResponse])
def get_pending_payments(
    db: Session = Depends(get_db),
    superadmin: User = Depends(get_current_superadmin)
):
    """
    Get all pending payment transactions.
    """
    try:
        service = PurchaseTransactionService(db)
        pending_transactions = service.get_pending_payments()
        
        result = []
        for transaction in pending_transactions:
            response = PurchaseTransactionResponse.model_validate(transaction)
            if transaction.business:
                response.company = transaction.business.business_name
            result.append(response)
        
        logger.info(f"Pending payments retrieved by {superadmin.email}: {len(result)} found")
        return result
        
    except Exception as e:
        logger.error(f"Error fetching pending payments: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error fetching pending payments: {str(e)}"
        )


@router.get("/overdue", response_model=List[PurchaseTransactionResponse])
def get_overdue_payments(
    db: Session = Depends(get_db),
    superadmin: User = Depends(get_current_superadmin)
):
    """
    Get overdue payment transactions.
    """
    try:
        service = PurchaseTransactionService(db)
        overdue_transactions = service.get_overdue_payments()
        
        result = []
        for transaction in overdue_transactions:
            response = PurchaseTransactionResponse.model_validate(transaction)
            if transaction.business:
                response.company = transaction.business.business_name
            result.append(response)
        
        logger.info(f"Overdue payments retrieved by {superadmin.email}: {len(result)} found")
        return result
        
    except Exception as e:
        logger.error(f"Error fetching overdue payments: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error fetching overdue payments: {str(e)}"
        )


@router.post("/{transaction_id}/process-payment", response_model=Dict[str, Any])
def process_payment(
    transaction_id: int,
    payment_method: str,
    payment_reference: str,
    gateway_response: Optional[str] = None,
    db: Session = Depends(get_db),
    superadmin: User = Depends(get_current_superadmin)
):
    """
    Process payment for a transaction.
    """
    try:
        service = PurchaseTransactionService(db)
        result = service.process_payment(
            transaction_id,
            payment_method,
            payment_reference,
            gateway_response
        )
        
        logger.info(f"Payment processed for transaction {transaction_id} by {superadmin.email}")
        return result
        
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    except Exception as e:
        logger.error(f"Error processing payment for transaction {transaction_id}: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error processing payment: {str(e)}"
        )


# ============================================================================
# EXPORT FUNCTIONALITY
# ============================================================================

@router.get("/export/excel")
def export_transactions_to_excel(
    status_filter: Optional[str] = Query(None, description="Filter by payment status"),
    payment_method: Optional[str] = Query(None, description="Filter by payment method"),
    date_from: Optional[datetime] = Query(None, description="Filter from date"),
    date_to: Optional[datetime] = Query(None, description="Filter to date"),
    search: Optional[str] = Query(None, description="Search term"),
    db: Session = Depends(get_db),
    superadmin: User = Depends(get_current_superadmin)
):
    """
    Export purchase transactions to Excel file.
    """
    try:
        service = PurchaseTransactionService(db)
        
        # Create filters object
        filters = TransactionFilters(
            status=status_filter,
            payment_method=payment_method,
            date_from=date_from,
            date_to=date_to,
            search_term=search,
            sort_by="transaction_date",
            sort_order="desc"
        )
        
        # Get all transactions (no limit for export)
        transactions = service.get_all_transactions_for_frontend(
            skip=0,
            limit=10000,  # Large limit for export
            filters=filters
        )
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Purchase Transactions"
        
        # Define headers
        headers = ["SN", "Invoice ID", "Company", "Email", "Date", "Amount (₹)", "Payment Method", "Status"]
        
        # Style for headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Write data
        for idx, transaction in enumerate(transactions, 1):
            ws.cell(row=idx + 1, column=1, value=idx)
            ws.cell(row=idx + 1, column=2, value=transaction.id)
            ws.cell(row=idx + 1, column=3, value=transaction.company)
            ws.cell(row=idx + 1, column=4, value=transaction.email)
            ws.cell(row=idx + 1, column=5, value=transaction.date)
            ws.cell(row=idx + 1, column=6, value=transaction.amount)
            ws.cell(row=idx + 1, column=7, value=transaction.method)
            ws.cell(row=idx + 1, column=8, value=transaction.status)
            
            # Apply status color coding
            status_cell = ws.cell(row=idx + 1, column=8)
            if transaction.status == "Paid":
                status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                status_cell.font = Font(color="006100")
            elif transaction.status == "Unpaid" or transaction.status == "Pending":
                status_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                status_cell.font = Font(color="9C6500")
            elif transaction.status == "Failed":
                status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                status_cell.font = Font(color="9C0006")
        
        # Adjust column widths
        column_widths = [8, 15, 35, 35, 15, 15, 20, 15]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width
        
        # Save to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Generate filename with timestamp
        filename = f"purchase_transactions_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        logger.info(f"Excel export generated by {superadmin.email}: {len(transactions)} transactions")
        
        return Response(
            content=excel_file.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )
        
    except Exception as e:
        logger.error(f"Error exporting transactions to Excel: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error exporting transactions: {str(e)}"
        )